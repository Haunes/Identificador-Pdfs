"""
EXTRACTOR DE TABLAS PDF A EXCEL - STREAMLIT CON DIBUJO INTERACTIVO
Dibuja rectángulos sobre el PDF para seleccionar tablas
Compatible con Streamlit Cloud
"""

import streamlit as st
from streamlit_drawable_canvas import st_canvas
import fitz  # PyMuPDF
import pandas as pd
from PIL import Image
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import camelot
import json

st.set_page_config(
    page_title="Extractor PDF a Excel",
    page_icon="📊",
    layout="wide"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .instruction-box {
        background-color: #e3f2fd;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1976d2;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# Inicializar session state
if 'pdf_doc' not in st.session_state:
    st.session_state.pdf_doc = None
if 'pdf_file' not in st.session_state:
    st.session_state.pdf_file = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = 0
if 'selected_areas' not in st.session_state:
    st.session_state.selected_areas = []
if 'canvas_key' not in st.session_state:
    st.session_state.canvas_key = 0

def load_pdf(uploaded_file):
    """Cargar PDF usando PyMuPDF"""
    try:
        pdf_bytes = uploaded_file.read()
        st.session_state.pdf_file = pdf_bytes
        st.session_state.pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        st.session_state.current_page = 0
        st.session_state.selected_areas = []
        st.session_state.canvas_key += 1
        return True
    except Exception as e:
        st.error(f"❌ Error al cargar PDF: {e}")
        return False

def render_page(page_num, zoom=2.0):
    """Renderizar página del PDF como imagen"""
    page = st.session_state.pdf_doc[page_num]
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return img, pix.width, pix.height

def extract_table_from_area(page_num, x0, y0, x1, y1, img_width, img_height):
    """Extraer tabla de un área específica"""
    try:
        # Guardar PDF temporalmente
        with open("temp_pdf.pdf", "wb") as f:
            f.write(st.session_state.pdf_file)
        
        # Obtener dimensiones de la página en puntos PDF
        page = st.session_state.pdf_doc[page_num]
        pdf_width = page.rect.width
        pdf_height = page.rect.height
        
        # Convertir coordenadas de imagen a coordenadas PDF
        scale_x = pdf_width / img_width
        scale_y = pdf_height / img_height
        
        pdf_x0 = x0 * scale_x
        pdf_y0 = y0 * scale_y
        pdf_x1 = x1 * scale_x
        pdf_y1 = y1 * scale_y
        
        # Camelot usa origen abajo-izquierda
        tabla_area = [
            str(pdf_x0),
            str(pdf_height - pdf_y1),
            str(pdf_x1),
            str(pdf_height - pdf_y0)
        ]
        
        # Extraer con Camelot
        tablas = camelot.read_pdf(
            "temp_pdf.pdf",
            pages=str(page_num + 1),
            flavor='stream',
            table_areas=[','.join(tabla_area)]
        )
        
        if len(tablas) > 0:
            return tablas[0].df
        return None
        
    except Exception as e:
        st.error(f"❌ Error extrayendo tabla: {e}")
        return None

def save_to_excel(dataframes, output_filename):
    """Guardar DataFrames a Excel"""
    try:
        combined_data = []
        for df in dataframes:
            if not df.empty:
                if combined_data:
                    separador = pd.DataFrame([['']*df.shape[1]])
                    combined_data.append(separador)
                combined_data.append(df)
        
        if not combined_data:
            return None
        
        df_final = pd.concat(combined_data, ignore_index=True)
        df_final = df_final.replace('', pd.NA).dropna(how='all').fillna('')
        
        # Guardar a Excel
        output = io.BytesIO()
        df_final.to_excel(output, index=False, engine='openpyxl')
        
        # Formatear
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        
        # Formatear encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        return excel_output, df_final
        
    except Exception as e:
        st.error(f"❌ Error guardando Excel: {e}")
        return None, None

# ============================================
# INTERFAZ PRINCIPAL
# ============================================

st.markdown('<p class="main-header">🎯 Extractor de Tablas PDF a Excel</p>', unsafe_allow_html=True)

# Columnas principales
col_sidebar, col_main = st.columns([1, 3])

with col_sidebar:
    st.markdown("""
    <div class="instruction-box">
    <b>📖 Instrucciones:</b><br><br>
    1️⃣ Sube tu PDF<br>
    2️⃣ <b>Dibuja rectángulos</b> sobre las tablas<br>
    3️⃣ Navega entre páginas si necesitas<br>
    4️⃣ Click en "Extraer a Excel"
    </div>
    """, unsafe_allow_html=True)
    
    # Upload PDF
    uploaded_file = st.file_uploader("📂 Cargar PDF", type=['pdf'])
    
    if uploaded_file:
        if st.session_state.pdf_doc is None or uploaded_file.name != getattr(st.session_state, 'last_file_name', ''):
            if load_pdf(uploaded_file):
                st.session_state.last_file_name = uploaded_file.name
                st.success(f"✅ {st.session_state.pdf_doc.page_count} página(s)")
    
    if st.session_state.pdf_doc:
        st.divider()
        
        # Navegación
        st.subheader("📄 Navegación")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col1:
            if st.button("◀", use_container_width=True, key="prev"):
                if st.session_state.current_page > 0:
                    st.session_state.current_page -= 1
                    st.session_state.canvas_key += 1
                    st.rerun()
        
        with col2:
            st.markdown(f"<center><b>Pág {st.session_state.current_page + 1}/{st.session_state.pdf_doc.page_count}</b></center>", unsafe_allow_html=True)
        
        with col3:
            if st.button("▶", use_container_width=True, key="next"):
                if st.session_state.current_page < st.session_state.pdf_doc.page_count - 1:
                    st.session_state.current_page += 1
                    st.session_state.canvas_key += 1
                    st.rerun()
        
        st.divider()
        
        # Áreas seleccionadas
        st.subheader("📦 Áreas Seleccionadas")
        
        if st.session_state.selected_areas:
            for i, area in enumerate(st.session_state.selected_areas):
                col_info, col_del = st.columns([4, 1])
                with col_info:
                    st.caption(f"**{i+1}.** Pág {area['page']+1}")
                with col_del:
                    if st.button("🗑️", key=f"del_{i}", use_container_width=True):
                        st.session_state.selected_areas.pop(i)
                        st.session_state.canvas_key += 1
                        st.rerun()
            
            st.divider()
            
            # Botón extraer
            if st.button("✅ **EXTRAER A EXCEL**", type="primary", use_container_width=True):
                with st.spinner("🔄 Extrayendo datos..."):
                    dataframes = []
                    
                    for area in st.session_state.selected_areas:
                        df = extract_table_from_area(
                            area['page'],
                            area['x0'], area['y0'],
                            area['x1'], area['y1'],
                            area['img_width'], area['img_height']
                        )
                        if df is not None and not df.empty:
                            dataframes.append(df)
                    
                    if dataframes:
                        excel_file, df_preview = save_to_excel(dataframes, "extraido.xlsx")
                        
                        if excel_file:
                            st.success(f"✅ **{len(df_preview)} filas extraídas!**")
                            
                            # Vista previa
                            with st.expander("👀 Vista previa de datos"):
                                st.dataframe(df_preview.head(20), use_container_width=True)
                            
                            # Botón de descarga
                            st.download_button(
                                label="⬇️ **Descargar Excel**",
                                data=excel_file,
                                file_name="pdf_extraido.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                    else:
                        st.error("❌ No se pudo extraer ningún dato")
            
            if st.button("🗑️ Limpiar Todo", use_container_width=True):
                st.session_state.selected_areas = []
                st.session_state.canvas_key += 1
                st.rerun()
        else:
            st.info("No hay áreas seleccionadas")

# Main area - Canvas para dibujar
with col_main:
    if st.session_state.pdf_doc:
        st.subheader(f"🖼️ Página {st.session_state.current_page + 1} - Dibuja rectángulos sobre las tablas")
        
        try:
            # Renderizar página
            img, img_width, img_height = render_page(st.session_state.current_page)
            
            # Crear objetos iniciales para áreas ya seleccionadas en esta página
            initial_drawing = {
                "version": "4.4.0",
                "objects": []
            }
            
            for area in st.session_state.selected_areas:
                if area['page'] == st.session_state.current_page:
                    initial_drawing["objects"].append({
                        "type": "rect",
                        "version": "4.4.0",
                        "left": area['x0'],
                        "top": area['y0'],
                        "width": area['x1'] - area['x0'],
                        "height": area['y1'] - area['y0'],
                        "fill": "rgba(255, 0, 0, 0.3)",
                        "stroke": "red",
                        "strokeWidth": 3
                    })
            
            # Canvas interactivo
            canvas_result = st_canvas(
                fill_color="rgba(255, 0, 0, 0.3)",
                stroke_width=3,
                stroke_color="red",
                background_image=img,
                update_streamlit=True,
                height=img_height,
                width=img_width,
                drawing_mode="rect",
                key=f"canvas_{st.session_state.canvas_key}",
                initial_drawing=initial_drawing,
                display_toolbar=True,
            )
            
            # Procesar nuevos rectángulos dibujados
            if canvas_result.json_data is not None:
                objects = canvas_result.json_data["objects"]
                
                # Filtrar solo rectángulos nuevos (que no están ya en selected_areas)
                new_rects = []
                for obj in objects:
                    if obj["type"] == "rect":
                        x0 = obj["left"]
                        y0 = obj["top"]
                        x1 = obj["left"] + obj["width"]
                        y1 = obj["top"] + obj["height"]
                        
                        # Verificar si ya existe
                        is_existing = False
                        for area in st.session_state.selected_areas:
                            if (area['page'] == st.session_state.current_page and
                                abs(area['x0'] - x0) < 5 and
                                abs(area['y0'] - y0) < 5):
                                is_existing = True
                                break
                        
                        if not is_existing and obj["width"] > 10 and obj["height"] > 10:
                            new_rects.append({
                                'page': st.session_state.current_page,
                                'x0': x0,
                                'y0': y0,
                                'x1': x1,
                                'y1': y1,
                                'img_width': img_width,
                                'img_height': img_height
                            })
                
                # Agregar nuevos rectángulos
                if new_rects:
                    st.session_state.selected_areas.extend(new_rects)
                    st.rerun()
            
            st.caption(f"📐 Dimensiones: {img_width} x {img_height} px")
            
        except Exception as e:
            st.error(f"❌ Error mostrando página: {e}")
    else:
        st.info("👈 **Carga un archivo PDF** desde el panel lateral para comenzar")
        
        st.markdown("""
        ### ✨ Características:
        
        - 🖱️ **Dibuja con el mouse** directamente sobre el PDF
        - 📄 **Múltiples páginas** - navega y selecciona en cualquier página
        - 📊 **Múltiples tablas** - selecciona cuantas necesites
        - ✅ **Exporta a Excel** con formato profesional
        - 🎨 **Vista previa** de las áreas seleccionadas
        
        ### 🚀 ¡Es muy fácil!
        
        1. Sube tu PDF
        2. Haz clic y arrastra sobre las tablas
        3. Descarga el Excel
        """)

st.divider()
st.caption("🔧 Extractor Visual de Tablas PDF | Compatible con Streamlit Cloud")